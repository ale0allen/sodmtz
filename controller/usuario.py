from openpyxl import Workbook, load_workbook
from collections import OrderedDict
import os

diretorio_atual = os.getcwd()
diretorio_pai = os.path.dirname(diretorio_atual)
diretorio_pai = diretorio_pai + '\\planilhas'

# construir um caminho relativo para um arquivo no diretório atual
caminho_rel = os.path.join(diretorio_pai, "Usuarios.xlsx")
filename = caminho_rel

wb = load_workbook(filename=filename)
ws = wb.active


from model.usuario import Usuario

class UsuarioController:
    def __init__(self, cpf_usuario=None, nome_usuario=None):
        if cpf_usuario is None:
            self.cpf_usuario = ""
        else:
            self.cpf_usuario = cpf_usuario
        if nome_usuario is None:
            self.nome_usuario = ""
        else:
            self.nome_usuario = nome_usuario

    def set_usuario(self):
        cpf_usuario = self.cpf_usuario
        nome_usuario = self.nome_usuario
        proxima_linha = ws.max_row + 1

        ws.cell(row=proxima_linha, column=1).value = proxima_linha
        ws.cell(row=proxima_linha, column=2).value = cpf_usuario
        ws.cell(row=proxima_linha, column=3).value = nome_usuario
        wb.save(filename=filename)

    def get_usuario(self):
        print(f'Esse é o cpf a ser buscado: {self.cpf_usuario}')
        for linha in range(2, ws.max_row + 1):
            if ws.cell(row=linha, column=2).value == int(self.cpf_usuario):
                self.nome_usuario = ws.cell(row=linha, column=3).value
                return self
        return None


    def update_usuario(self, novo_nome):
        print(f'Esse é o cpf: {self.cpf_usuario}')
        for linha in range(2, ws.max_row + 1):
            if ws.cell(row=linha, column=2).value == int(self.cpf_usuario):
                ws.cell(row=linha, column=3).value = novo_nome
                wb.save(filename=filename)
        print('Usuario Atualizado!')

    def todos_usuarios(self):
        todas_linhas = []
        for linha in range(2, ws.max_row + 1):
            cod = ws.cell(row=linha, column=1).value
            cpf = ws.cell(row=linha, column=2).value
            nome_usuario = ws.cell(row=linha, column=3).value
            user = UsuarioController(cpf_usuario=cpf, nome_usuario=nome_usuario)
            todas_linhas.append(user)
        return todas_linhas
