from openpyxl import Workbook, load_workbook
from collections import OrderedDict
import os


diretorio_atual = os.getcwd()
diretorio_pai = os.path.dirname(diretorio_atual)
diretorio_pai = diretorio_pai + '\\planilhas'

# construir um caminho relativo para um arquivo no diretório atual
caminho_rel = os.path.join(diretorio_pai, "Perfis.xlsx")
filename = caminho_rel

wb = load_workbook(filename=filename)
ws = wb.active


from model.perfil import Perfil

class PerfilController:

    def __init__(self, codigo_perfil=None, nome_perfil=None, nome_sistema=None, descricao_perfil=None ):
        if codigo_perfil is None:
            self.codigo_perfil = ""
        else:
            self.codigo_perfil = codigo_perfil
        if nome_perfil is None:
            self.nome_perfil = ""
        else:
            self.nome_perfil = nome_perfil
        if nome_sistema is None:
            self.nome_sistema = ""
        else:
            self.nome_sistema = nome_sistema
        if descricao_perfil is None:
            self.descricao_perfil = ""
        else:
            self.descricao_perfil = descricao_perfil

    def set_perfil(self):
        nome_sistema = self.nome_sistema
        nome_perfil = self.nome_perfil
        descricao_perfil = self.descricao_perfil

        proxima_linha = ws.max_row + 1
        ws.cell(row=proxima_linha, column=1).value = proxima_linha
        ws.cell(row=proxima_linha, column=2).value = nome_perfil
        ws.cell(row=proxima_linha, column=3).value = nome_sistema
        ws.cell(row=proxima_linha, column=4).value = descricao_perfil
        wb.save(filename=filename)

    def get_perfil(self):
        print(f'Esse é o codigo a ser buscado: {self.codigo_perfil}')
        for linha in range(2, ws.max_row + 1):
            if ws.cell(row=linha, column=1).value == self.codigo_perfil:
                self.nome_perfil = ws.cell(row=linha, column=2).value
                self.nome_sistema = ws.cell(row=linha, column=3).value
                self.descricao_perfil = ws.cell(row=linha, column=4).value
                return self
        print('Sistema não encontrado')
        return None

    def todos_perfis(self):
        todas_linhas = []
        for linha in range(2, ws.max_row + 1):
            codigo = ws.cell(row=linha, column=1).value
            nome_perfil = ws.cell(row=linha, column=2).value
            nome_sistema = ws.cell(row=linha, column=3).value
            descricao_perfil = ws.cell(row=linha, column=4).value
            perf = Perfil(codigo, nome_perfil, nome_sistema, descricao_perfil)
            todas_linhas.append(perf)
        return todas_linhas



    def __str__(self):
        return f'nome: {self.nome_perfil} - sistema: {self.nome_sistema} - descrição: {self.descricao_perfil}'
