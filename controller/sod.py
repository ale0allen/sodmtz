from openpyxl import Workbook, load_workbook
from collections import OrderedDict
import os

diretorio_atual = os.getcwd()
diretorio_pai = os.path.dirname(diretorio_atual)
diretorio_pai = diretorio_pai + '\\planilhas'

# construir um caminho relativo para um arquivo no diretório atual
caminho_rel = os.path.join(diretorio_pai, "MatrizSod.xlsx")
filename = caminho_rel


wb = load_workbook(filename=filename)
ws = wb.active


from model.sod import Sod


class SodController:

    def __init__(self, codigo=None, codigo_perfil1=None, nome_perfil1=None, sistema_perfil1=None,
                 codigo_perfil2=None, nome_perfil2=None, sistema_perfil2=None):

        if codigo is None:
            self.codigo = ""
        else:
            self.codigo = codigo

        if codigo_perfil1 is None:
            self.codigo_perfil1 = ""
        else:
            self.codigo_perfil1 = codigo_perfil1

        if nome_perfil1 is None:
            self.nome_perfil1 = ""
        else:
            self.nome_perfil1 = nome_perfil1

        if sistema_perfil1 is None:
            self.sistema_perfil1 = ""
        else:
            self.sistema_perfil1 = sistema_perfil1

        if codigo_perfil2 is None:
            self.codigo_perfil2 = ""
        else:
            self.codigo_perfil2 = codigo_perfil2

        if nome_perfil2 is None:
            self.nome_perfil2 = ""
        else:
            self.nome_perfil2 = nome_perfil2

        if sistema_perfil2 is None:
            self.sistema_perfil2 = ""
        else:
            self.sistema_perfil2 = sistema_perfil2

    def set_sod(self):
        codigo_perfil1 = self.codigo_perfil1
        nome_perfil1 = self.nome_perfil1
        sistema_perfil1 = self.sistema_perfil1
        codigo_perfil2 = self.codigo_perfil2
        nome_perfil2 = self.nome_perfil2
        sistema_perfil2 = self.sistema_perfil2

        proxima_linha = ws.max_row + 1
        ws.cell(row=proxima_linha, column=1).value = proxima_linha
        ws.cell(row=proxima_linha, column=2).value = codigo_perfil1
        ws.cell(row=proxima_linha, column=3).value = nome_perfil1
        ws.cell(row=proxima_linha, column=4).value = sistema_perfil1
        ws.cell(row=proxima_linha, column=5).value = codigo_perfil2
        ws.cell(row=proxima_linha, column=6).value = nome_perfil2
        ws.cell(row=proxima_linha, column=7).value = sistema_perfil2
        wb.save(filename=filename)

    def todos_sod(self):
        todas_linhas = []
        for linha in range(2, ws.max_row + 1):
            codigo = ws.cell(row=linha, column=1).value
            codigo_perfil1 = ws.cell(row=linha, column=2).value
            nome_perfil1 = ws.cell(row=linha, column=3).value
            sistema_perfil1 = ws.cell(row=linha, column=4).value
            codigo_perfil2 = ws.cell(row=linha, column=5).value
            nome_perfil2 = ws.cell(row=linha, column=6).value
            sistema_perfil2 = ws.cell(row=linha, column=7).value
            sod = Sod(codigo,codigo_perfil1, nome_perfil1, sistema_perfil1, codigo_perfil2, nome_perfil2, sistema_perfil2)
            todas_linhas.append(sod)
        return todas_linhas

    def get_sod(self):
        print(f'Esse é o codigo a ser buscado: {self.codigo}')
        for linha in range(2, ws.max_row + 1):
            if ws.cell(row=linha, column=1).value == self.codigo:
                self.codigo_perfil1 = ws.cell(row=linha, column=2).value
                self.nome_perfil1 = ws.cell(row=linha, column=3).value
                self.sistema_perfil1 = ws.cell(row=linha, column=4).value
                self.codigo_perfil2 = ws.cell(row=linha, column=5).value
                self.nome_perfil2 = ws.cell(row=linha, column=6).value
                self.sistema_perfil2 = ws.cell(row=linha, column=7).value
                return self
        print('Sistema não encontrado')
        return None

    def update_sod(self, novo_codigo_perfil1, novo_nome_perfil1, novo_sistema_perfil1, novo_codigo_perfil2,
                   novo_nome_perfil2, novo_sistema_perfil2):
        for linha in range(2, ws.max_row + 1):
            if ws.cell(row=linha, column=1).value == self.codigo:
                ws.cell(row=linha, column=2).value = novo_codigo_perfil1
                ws.cell(row=linha, column=3).value = novo_nome_perfil1
                ws.cell(row=linha, column=4).value = novo_sistema_perfil1
                ws.cell(row=linha, column=5).value = novo_codigo_perfil2
                ws.cell(row=linha, column=6).value = novo_nome_perfil2
                ws.cell(row=linha, column=7).value = novo_sistema_perfil2
                wb.save(filename=filename)
        print('Sistema Atualizado!')

    def delete_sod(self):
        for linha in range(2, ws.max_row + 1):
            if ws.cell(row=linha, column=1).value == self.codigo:
                ws.delete_rows(linha)
                wb.save(filename=filename)

