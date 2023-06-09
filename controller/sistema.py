from openpyxl import Workbook, load_workbook
from collections import OrderedDict
import os

diretorio_atual = os.getcwd()
diretorio_pai = os.path.dirname(diretorio_atual)
diretorio_pai = diretorio_pai + '\\planilhas'

# construir um caminho relativo para um arquivo no diretório atual
caminho_rel = os.path.join(diretorio_pai, "Sistemas.xlsx")
filename = caminho_rel



wb = load_workbook(filename=filename)
ws = wb.active

from model.sistema import Sistema


class SistemaController:

    def __init__(self, codigo=None, nome=None):
        if codigo is None:
            self.codigo = ""
        else:
            self.codigo = codigo
        if nome is None:
            self.nome = ""
        else:
            self.nome = nome

    def get_sistema(self):
        print(f'Esse é o codigo a ser buscado: {self.codigo}')
        for linha in range(2, ws.max_row + 1):
            if ws.cell(row=linha, column=1).value == self.codigo:
                self.nome = ws.cell(row=linha, column=2).value
                return  self
        print('Sistema não encontrado')
        return None

    def update_sistema(self, novo_nome):
        if self.get_sistema():
            for linha in range(2, ws.max_row + 1):
                if ws.cell(row=linha, column=1).value == self.codigo:
                    ws.cell(row=linha, column=2).value = novo_nome
                    wb.save(filename=filename)
            print('Sistema Atualizado!')
        else:
            print('Sistema não encontrado, impossivel atualizar')

    def delete_sistema(self):
        if self.get_sistema():
            for linha in range(2, ws.max_row + 1):
                if ws.cell(row=linha, column=1).value == self.codigo:
                    # ws.cell(row=linha, column=3).value = 1
                    ws.delete_rows(linha)
                    wb.save(filename=filename)
            print('Sistema Deletado!')
        else:
            print('Sistema não encontrado, impossivel deletar')

    def todos_sistemas(self):
        todas_linhas = []
        for linha in range(2, ws.max_row + 1):
            codigo = ws.cell(row=linha, column=1).value
            nome_sistema = ws.cell(row=linha, column=2).value
            sis = Sistema(codigo,nome_sistema)
            todas_linhas.append(sis)
        return todas_linhas

    def set_sistema(self):
        if self.get_sistema():
            print('Sistema ja cadastrado!')
        else:
            nome = self.nome
            proxima_linha = ws.max_row + 1
            cod = "{:015}".format(proxima_linha)
            ws.cell(row=proxima_linha, column=1).value = proxima_linha
            ws.cell(row=proxima_linha, column=2).value = nome
            wb.save(filename=filename)
    def __str__(self):
        return f'Código => {self.codigo} Nome => {self.nome}'



