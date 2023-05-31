from openpyxl import Workbook, load_workbook
from collections import OrderedDict
import os
from controller.sod import SodController

diretorio_atual = os.getcwd()
diretorio_pai = os.path.dirname(diretorio_atual)
diretorio_pai = diretorio_pai + '\\planilhas'

# construir um caminho relativo para um arquivo no diretório atual
caminho_rel = os.path.join(diretorio_pai, "Vinculos.xlsx")
filename = caminho_rel

wb = load_workbook(filename=filename)
ws = wb.active


class VinculoController:
    def __init__(self, cpf_usuario=None, id_perfil=None):
        if cpf_usuario is None:
            self.cpf_usuario = ''
        else:
            self.cpf_usuario = cpf_usuario
        if id_perfil is None:
            self.id_perfil = ''
        else:
            self.id_perfil = id_perfil


    def set_vinculo(self):
        vericacao = self.verifica_sod()
        print(vericacao)

        if vericacao is False or vericacao is None:
            cpf_usuario = self.cpf_usuario
            id_perfil = self.id_perfil
            proxima_linha = ws.max_row + 1

            ws.cell(row=proxima_linha, column=1).value = proxima_linha
            ws.cell(row=proxima_linha, column=2).value = cpf_usuario
            ws.cell(row=proxima_linha, column=3).value = id_perfil
            wb.save(filename=filename)
            print('Novo perfil criado')
        else:
            print('Conflito na Matriz... não é possivel cadastrar')


    def get_vinculos_cpf(self):
        cpf_usuario = self.cpf_usuario
        todas_linhas = []

        for linha in range(2, ws.max_row + 1):
            if ws.cell(row=linha, column=2).value == cpf_usuario:
                cod = linha
                id_perfil = ws.cell(row=linha, column=3)
                vinculo = VinculoController(cpf_usuario=cpf_usuario, id_perfil=id_perfil)
                todas_linhas.append(vinculo)
        return todas_linhas

    def verifica_sod(self):
        perfil_novo = self.id_perfil
        matriz = SodController
        matriz_sod = matriz.todos_sod(self)
        lista = []
        for matriz in matriz_sod:
            mt1 = matriz.codigo_perfil1
            mt2 = matriz.codigo_perfil2
            lista.append([mt1, mt2])

        cadastrados = self.get_vinculos_cpf()
        if cadastrados is not None:
            for cadastro in cadastrados:
                perfil_cadastrado = cadastro.id_perfil.value

                conflito = False  # Variável para acompanhar se há um conflito
                for linha_matriz in lista:
                    perfil1 = linha_matriz[0]
                    perfil2 = linha_matriz[1]

                    if perfil_cadastrado == perfil1 and perfil_novo == perfil2 or \
                            perfil_novo == perfil1 and perfil_cadastrado == perfil2:
                        print('Conflito na matriz')
                        conflito = True  # Atualiza a variável conflito para True

                if conflito:
                    return True

        return False










